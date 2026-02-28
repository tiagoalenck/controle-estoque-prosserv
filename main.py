import os
import re
import sqlite3
import traceback
import hashlib
import hmac
import calendar
import csv
import io
import tempfile
import xml.etree.ElementTree as ET
from datetime import datetime, date, timedelta
from pathlib import Path
from urllib.parse import urlencode

from fastapi import FastAPI, Request, Form
from fastapi.responses import HTMLResponse, RedirectResponse, PlainTextResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from fastapi.staticfiles import StaticFiles

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ✅ Ambiente (dev | prod)
ENV = os.getenv("ENV", "dev").lower()

app = FastAPI()

BASE_DIR = Path(__file__).resolve().parent
TEMPLATES_DIR = BASE_DIR / "templates"

# ============================================================
# ✅ BANCO (SQLite)
# - NUNCA suba estoque.db pro GitHub
# - Use DB_PATH pra definir onde salvar o banco (prod / servidor / etc.)
# - Se não tiver DB_PATH, salva em ./data/estoque.db
# ============================================================
DB_PATH_ENV = os.getenv("DB_PATH", "").strip()

if DB_PATH_ENV:
    DB_FILE = Path(DB_PATH_ENV).expanduser().resolve()
else:
    # Se você já tinha um estoque.db antigo na raiz, mantém pra não quebrar
    legacy = BASE_DIR / "estoque.db"
    if legacy.exists():
        DB_FILE = legacy
    else:
        DB_FILE = BASE_DIR / "data" / "estoque.db"

DB_FILE.parent.mkdir(parents=True, exist_ok=True)
DB = str(DB_FILE)

# ✅ STATIC: para logo funcionar em /static/logo.png (e também /static/styles.css)
STATIC_DIR = BASE_DIR / "static"
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
templates.env.auto_reload = True
templates.env.cache = {}
templates.env.globals["os"] = os

# ✅ Sessão (login) - segura
SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    SECRET_KEY = "dev-only-change-me"
    if ENV == "prod":
        raise RuntimeError("ENV=prod exige SECRET_KEY definida no ambiente.")

app.add_middleware(
    SessionMiddleware,
    secret_key=SECRET_KEY,
    same_site="lax",
    https_only=(ENV == "prod"),
)

@app.exception_handler(Exception)
async def show_exceptions(request: Request, exc: Exception):
    if ENV == "prod":
        return PlainTextResponse("Erro interno. Contate o administrador.", status_code=500)
    return PlainTextResponse(traceback.format_exc(), status_code=500)

def now_txt() -> str:
    return datetime.now().isoformat(timespec="seconds")

def _con():
    con = sqlite3.connect(DB, check_same_thread=False)
    con.execute("PRAGMA foreign_keys=ON")
    con.execute("PRAGMA journal_mode=WAL")
    return con

# ==========================
# VALIDADE (PADRONIZAÇÃO)
# ==========================
def _parse_validade_to_date(s: str):
    """
    Aceita:
      - YYYY-MM-DD
      - YYYY-MM
      - MM/YYYY
      - DD/MM/YYYY
    Retorna date ou None.
    """
    s = (s or "").strip()
    if not s:
        return None

    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if m:
        y, mo, d = map(int, m.groups())
        return date(y, mo, d)

    m = re.match(r"^(\d{4})-(\d{2})$", s)
    if m:
        y, mo = map(int, m.groups())
        last_day = calendar.monthrange(y, mo)[1]
        return date(y, mo, last_day)

    m = re.match(r"^(\d{2})/(\d{4})$", s)
    if m:
        mo, y = map(int, m.groups())
        last_day = calendar.monthrange(y, mo)[1]
        return date(y, mo, last_day)

    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", s)
    if m:
        d, mo, y = map(int, m.groups())
        return date(y, mo, d)

    return None

def normalize_validade_ddmmyyyy(s: str) -> str:
    """
    Padroniza SEMPRE para DD/MM/YYYY para salvar no banco.
    Se não conseguir interpretar, levanta ValueError.
    """
    v = _parse_validade_to_date(s)
    if v is None:
        raise ValueError("Validade inválida. Use DD/MM/YYYY (ou: YYYY-MM-DD, YYYY-MM, MM/YYYY).")
    return v.strftime("%d/%m/%Y")

# ==========================
# AUTH (senha segura)
# ==========================
def _pw_hash(password: str, salt: bytes) -> bytes:
    return hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 120_000)

def make_password(password: str) -> str:
    salt = os.urandom(16)
    ph = _pw_hash(password, salt)
    return salt.hex() + ":" + ph.hex()

def verify_password(password: str, stored: str) -> bool:
    try:
        salt_hex, hash_hex = stored.split(":")
        salt = bytes.fromhex(salt_hex)
        expected = bytes.fromhex(hash_hex)
        got = _pw_hash(password, salt)
        return hmac.compare_digest(got, expected)
    except Exception:
        return False

def current_user(request: Request):
    return request.session.get("user")

def require_role(request: Request, roles: set[str]):
    u = current_user(request)
    if not u:
        raise PermissionError("Você precisa fazer login.")
    if u.get("role") not in roles:
        raise PermissionError("Sem permissão para acessar.")

# ==========================
# DB INIT + MIGRAÇÕES
# ==========================
def db_init():
    con = _con()
    cur = con.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS entradas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT NOT NULL,
            descricao TEXT NOT NULL,
            lote TEXT NOT NULL,
            laboratorio TEXT NOT NULL,
            validade TEXT NOT NULL,
            quantidade INTEGER NOT NULL,
            endereco TEXT NOT NULL,
            criado_em TEXT,
            criado_por TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS estoque (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT NOT NULL,
            descricao TEXT NOT NULL,
            lote TEXT NOT NULL,
            laboratorio TEXT NOT NULL,
            validade TEXT NOT NULL,
            endereco TEXT NOT NULL,
            quantidade INTEGER NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS separacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pedido TEXT NOT NULL UNIQUE,
            criado_em TEXT NOT NULL,
            status TEXT NOT NULL,
            cliente TEXT,
            separador TEXT,
            criado_por TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS separacao_itens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pedido TEXT NOT NULL,
            estoque_id INTEGER NOT NULL,
            codigo TEXT NOT NULL,
            descricao TEXT NOT NULL,
            lote TEXT NOT NULL,
            laboratorio TEXT NOT NULL,
            validade TEXT NOT NULL,
            endereco TEXT NOT NULL,
            quantidade INTEGER NOT NULL,
            criado_em TEXT NOT NULL,
            criado_por TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            criado_em TEXT NOT NULL
        )
    """)

    # ✅ migrações seguras (caso banco antigo)
    for sql in [
        "ALTER TABLE separacoes ADD COLUMN cliente TEXT",
        "ALTER TABLE separacoes ADD COLUMN separador TEXT",
        "ALTER TABLE separacoes ADD COLUMN criado_por TEXT",
        "ALTER TABLE entradas ADD COLUMN criado_por TEXT",
        "ALTER TABLE separacao_itens ADD COLUMN criado_por TEXT",
        "ALTER TABLE entradas ADD COLUMN laboratorio TEXT",
        "ALTER TABLE estoque ADD COLUMN laboratorio TEXT",
        "ALTER TABLE separacao_itens ADD COLUMN laboratorio TEXT",
    ]:
        try:
            cur.execute(sql)
        except Exception:
            pass

    # ✅ índices
    try:
        cur.execute("CREATE INDEX IF NOT EXISTS idx_estoque_busca ON estoque(codigo, descricao, lote, laboratorio, validade, endereco)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_itens_pedido ON separacao_itens(pedido)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_entradas_busca ON entradas(codigo, descricao, lote, laboratorio, validade, endereco)")
    except Exception:
        pass

    con.commit()

    # ✅ cria admin padrão se não existir nenhum user
    cur.execute("SELECT COUNT(*) FROM users")
    if int(cur.fetchone()[0]) == 0:
        cur.execute(
            "INSERT INTO users (username, password_hash, role, criado_em) VALUES (?, ?, ?, ?)",
            ("admin", make_password("admin123"), "admin", now_txt())
        )
        con.commit()

    con.close()

db_init()

# ==========================
# USERS
# ==========================
def user_get(username: str):
    con = _con()
    cur = con.cursor()
    cur.execute("SELECT username, password_hash, role FROM users WHERE username=? LIMIT 1", (username,))
    r = cur.fetchone()
    con.close()
    return r

def users_list():
    con = _con()
    cur = con.cursor()
    cur.execute("""
        SELECT id, username, role, criado_em
        FROM users
        ORDER BY role, username
    """)
    rows = cur.fetchall()
    con.close()
    return rows

def user_create(username: str, password: str, role: str):
    username = (username or "").strip()
    password = (password or "").strip()
    role = (role or "").strip()

    if not username:
        raise ValueError("Informe o usuário.")
    if len(username) < 3:
        raise ValueError("Usuário muito curto (mínimo 3 caracteres).")
    if not password or len(password) < 4:
        raise ValueError("Senha muito curta (mínimo 4 caracteres).")
    if role not in {"admin", "recebimento", "separador"}:
        raise ValueError("Perfil inválido.")

    con = _con()
    cur = con.cursor()
    try:
        cur.execute(
            "INSERT INTO users (username, password_hash, role, criado_em) VALUES (?, ?, ?, ?)",
            (username, make_password(password), role, now_txt())
        )
        con.commit()
    except sqlite3.IntegrityError:
        raise ValueError("Esse usuário já existe.")
    finally:
        con.close()

def _admin_count():
    con = _con()
    cur = con.cursor()
    cur.execute("SELECT COUNT(*) FROM users WHERE role='admin'")
    n = int(cur.fetchone()[0])
    con.close()
    return n

def user_delete(user_id: int, current_username: str):
    con = _con()
    cur = con.cursor()

    cur.execute("SELECT username, role FROM users WHERE id=? LIMIT 1", (int(user_id),))
    r = cur.fetchone()
    if not r:
        con.close()
        raise ValueError("Usuário não encontrado.")

    username, role = r

    if username == current_username:
        con.close()
        raise ValueError("Você não pode excluir o seu próprio usuário.")

    if role == "admin" and _admin_count() <= 1:
        con.close()
        raise ValueError("Não é permitido excluir o último admin do sistema.")

    cur.execute("DELETE FROM users WHERE id=?", (int(user_id),))
    con.commit()
    con.close()

# ==========================
# ENTRADAS / ESTOQUE
# ==========================
def entrada_insert(codigo, descricao, lote, laboratorio, validade, quantidade, endereco, criado_por=""):
    con = _con()
    cur = con.cursor()
    cur.execute("""
        INSERT INTO entradas (codigo, descricao, lote, laboratorio, validade, quantidade, endereco, criado_em, criado_por)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (codigo, descricao, lote, laboratorio, validade, int(quantidade), endereco, now_txt(), criado_por))
    con.commit()
    con.close()

def entrada_delete(entrada_id: int):
    con = _con()
    cur = con.cursor()
    cur.execute("DELETE FROM entradas WHERE id=?", (int(entrada_id),))
    con.commit()
    con.close()

def estoque_upsert(codigo, descricao, lote, laboratorio, validade, endereco, qtd):
    con = _con()
    cur = con.cursor()
    cur.execute("""
        SELECT id, quantidade FROM estoque
        WHERE codigo=? AND descricao=? AND lote=? AND laboratorio=? AND validade=? AND endereco=?
        LIMIT 1
    """, (codigo, descricao, lote, laboratorio, validade, endereco))
    row = cur.fetchone()

    if row:
        estoque_id, qtd_atual = row
        cur.execute("UPDATE estoque SET quantidade=? WHERE id=?",
                    (int(qtd_atual) + int(qtd), int(estoque_id)))
    else:
        cur.execute("""
            INSERT INTO estoque (codigo, descricao, lote, laboratorio, validade, endereco, quantidade)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (codigo, descricao, lote, laboratorio, validade, endereco, int(qtd)))

    con.commit()
    con.close()

def estoque_list(q: str = "", limit=500):
    q = (q or "").strip()
    con = _con()
    cur = con.cursor()

    if q:
        like = f"%{q}%"
        cur.execute("""
            SELECT id, codigo, descricao, lote, validade, endereco, quantidade
            FROM estoque
            WHERE quantidade > 0
              AND (codigo LIKE ? OR descricao LIKE ? OR lote LIKE ? OR laboratorio LIKE ? OR validade LIKE ? OR endereco LIKE ?)
            ORDER BY descricao, validade, lote, endereco
            LIMIT ?
        """, (like, like, like, like, like, like, int(limit)))
    else:
        cur.execute("""
            SELECT id, codigo, descricao, lote, validade, endereco, quantidade
            FROM estoque
            WHERE quantidade > 0
            ORDER BY descricao, validade, lote, endereco
            LIMIT ?
        """, (int(limit),))

    rows = cur.fetchall()
    con.close()
    return rows

# ==========================
# SEPARAÇÃO
# ==========================
def separacao_get_or_create(pedido: str, cliente: str = "", separador: str = "", criado_por: str = ""):
    pedido = (pedido or "").strip()
    cliente = (cliente or "").strip()
    separador = (separador or "").strip()
    if not pedido:
        return

    con = _con()
    cur = con.cursor()

    cur.execute("SELECT id, cliente, separador, criado_por FROM separacoes WHERE pedido=? LIMIT 1", (pedido,))
    r = cur.fetchone()

    if not r:
        cur.execute("""
            INSERT INTO separacoes (pedido, criado_em, status, cliente, separador, criado_por)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (pedido, now_txt(), "ABERTO", cliente, separador, criado_por))
        con.commit()
    else:
        sep_id, cli_db, sep_db, criador_db = r
        novo_cliente = cliente if cliente else (cli_db or "")
        novo_separador = separador if separador else (sep_db or "")
        novo_criador = criador_db or criado_por

        cur.execute("""
            UPDATE separacoes
            SET cliente=?, separador=?, criado_por=?
            WHERE id=?
        """, (novo_cliente, novo_separador, novo_criador, int(sep_id)))
        con.commit()

    con.close()

def separacao_meta_get(pedido: str):
    pedido = (pedido or "").strip()
    if not pedido:
        return ("", "")

    con = _con()
    cur = con.cursor()
    cur.execute("SELECT cliente, separador FROM separacoes WHERE pedido=? LIMIT 1", (pedido,))
    r = cur.fetchone()
    con.close()

    if not r:
        return ("", "")
    return (r[0] or "", r[1] or "")

def separacao_status_get(pedido: str) -> str:
    pedido = (pedido or "").strip()
    if not pedido:
        return ""
    con = _con()
    cur = con.cursor()
    cur.execute("SELECT status FROM separacoes WHERE pedido=? LIMIT 1", (pedido,))
    r = cur.fetchone()
    con.close()
    return (r[0] if r else "") or ""

def separacao_set_status(pedido: str, status: str):
    pedido = (pedido or "").strip()
    status = (status or "").strip().upper()
    if not pedido:
        raise ValueError("Pedido inválido.")
    if status not in {"ABERTO", "FINALIZADO"}:
        raise ValueError("Status inválido.")

    con = _con()
    cur = con.cursor()
    cur.execute("UPDATE separacoes SET status=? WHERE pedido=?", (status, pedido))
    con.commit()
    con.close()

def separacao_itens_list(pedido: str):
    con = _con()
    cur = con.cursor()
    cur.execute("""
        SELECT id, estoque_id, codigo, descricao, lote, validade, endereco, quantidade, criado_em
        FROM separacao_itens
        WHERE pedido=?
        ORDER BY id DESC
    """, (pedido,))
    rows = cur.fetchall()
    con.close()
    return rows

def separacao_item_add(pedido: str, estoque_id: int, qtd: int, criado_por: str = ""):
    pedido = (pedido or "").strip()
    if not pedido:
        raise ValueError("Informe o número do pedido.")

    qtd = int(qtd)
    if qtd <= 0:
        raise ValueError("Quantidade inválida.")

    con = _con()
    cur = con.cursor()

    try:
        cur.execute("""
            SELECT id, codigo, descricao, lote, laboratorio, validade, endereco, quantidade
            FROM estoque
            WHERE id=?
            LIMIT 1
        """, (int(estoque_id),))
        row = cur.fetchone()
        if not row:
            raise ValueError("Item de estoque não encontrado.")

        _, codigo, descricao, lote, laboratorio, validade, endereco, disponivel = row

        cur.execute("""
            UPDATE estoque
               SET quantidade = quantidade - ?
             WHERE id = ?
               AND quantidade >= ?
        """, (qtd, int(estoque_id), qtd))

        if cur.rowcount != 1:
            raise ValueError(f"Qtd maior que disponível (disp: {disponivel}).")

        cur.execute("""
            INSERT INTO separacao_itens (pedido, estoque_id, codigo, descricao, lote, laboratorio, validade, endereco, quantidade, criado_em, criado_por)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (pedido, int(estoque_id), codigo, descricao, lote, laboratorio, validade, endereco, qtd, now_txt(), criado_por))

        con.commit()
    except Exception:
        con.rollback()
        raise
    finally:
        con.close()

def separacao_item_remove(item_id: int):
    con = _con()
    cur = con.cursor()

    try:
        cur.execute("""
            SELECT estoque_id, quantidade, pedido
            FROM separacao_itens
            WHERE id=?
            LIMIT 1
        """, (int(item_id),))
        r = cur.fetchone()
        if not r:
            raise ValueError("Item não encontrado.")

        estoque_id, qtd, pedido = r

        cur.execute("DELETE FROM separacao_itens WHERE id=?", (int(item_id),))
        cur.execute("UPDATE estoque SET quantidade = quantidade + ? WHERE id=?", (int(qtd), int(estoque_id)))

        con.commit()
        return pedido
    except Exception:
        con.rollback()
        raise
    finally:
        con.close()

# ==========================
# RELATÓRIO VALIDADE
# ==========================
def estoque_list_all_qty_gt0():
    con = _con()
    cur = con.cursor()
    cur.execute("""
        SELECT id, codigo, descricao, lote, validade, endereco, quantidade
        FROM estoque
        WHERE quantidade > 0
        ORDER BY descricao, validade, lote, endereco
    """)
    rows = cur.fetchall()
    con.close()
    return rows

def relatorio_validade(dias: int | None):
    hoje = date.today()
    dias = 90 if dias is None else int(dias)

    rows = estoque_list_all_qty_gt0()
    itens = []

    if dias == 0:
        for r in rows:
            estoque_id, codigo, descricao, lote, validade_txt, endereco, qtd = r
            vdate = _parse_validade_to_date(validade_txt)
            if vdate is None:
                continue
            dias_para = (vdate - hoje).days
            if vdate < hoje:
                itens.append((estoque_id, codigo, descricao, lote, validade_txt, endereco, qtd, dias_para))
    else:
        limite = hoje + timedelta(days=dias)
        for r in rows:
            estoque_id, codigo, descricao, lote, validade_txt, endereco, qtd = r
            vdate = _parse_validade_to_date(validade_txt)
            if vdate is None:
                continue
            dias_para = (vdate - hoje).days
            if hoje <= vdate <= limite:
                itens.append((estoque_id, codigo, descricao, lote, validade_txt, endereco, qtd, dias_para))

    itens.sort(key=lambda x: x[7])
    return itens

# ==========================
# IMPORTAR NF-e (XML)
# ==========================
def parse_nfe_xml(xml_bytes: bytes):
    root = ET.fromstring(xml_bytes)

    m = re.match(r"^\{(.+)\}", root.tag)
    ns_uri = m.group(1) if m else ""
    ns = {"nfe": ns_uri} if ns_uri else {}

    inf = root.find(".//nfe:infNFe", ns) if ns_uri else root.find(".//infNFe")
    if inf is None:
        raise ValueError("Não encontrei infNFe. Esse XML parece não ser uma NF-e válida.")

    dets = inf.findall("nfe:det", ns) if ns_uri else inf.findall("det")
    itens = []

    for det in dets:
        prod = det.find("nfe:prod", ns) if ns_uri else det.find("prod")
        if prod is None:
            continue

        def get(tag):
            el = prod.find(f"nfe:{tag}", ns) if ns_uri else prod.find(tag)
            return (el.text or "").strip() if el is not None else ""

        codigo = get("cProd")
        descricao = get("xProd")
        qtd_txt = get("qCom") or get("qTrib") or "0"

        try:
            qtd = int(float(qtd_txt.replace(",", ".")))
        except Exception:
            qtd = 0

        laboratorio = get("xFab") or "N/I"

        lote = ""
        validade = ""

        rastros = prod.findall("nfe:rastro", ns) if ns_uri else prod.findall("rastro")
        if rastros:
            r0 = rastros[0]
            nLote = r0.find("nfe:nLote", ns) if ns_uri else r0.find("nLote")
            dVal = r0.find("nfe:dVal", ns) if ns_uri else r0.find("dVal")
            lote = (nLote.text or "").strip() if nLote is not None else ""
            validade = (dVal.text or "").strip() if dVal is not None else ""

        if codigo and descricao and qtd > 0:
            itens.append({
                "codigo": codigo,
                "descricao": descricao,
                "quantidade": qtd,
                "lote": lote,
                "validade": validade,
                "laboratorio": laboratorio,
            })

    if not itens:
        raise ValueError("Não encontrei itens (det/prod) no XML.")
    return itens

# ==========================
# ROTAS NF-e
# ==========================
@app.get("/nfe/importar", response_class=HTMLResponse)
def nfe_importar_get(request: Request, ok: str = "", error: str = ""):
    if not current_user(request):
        return RedirectResponse(url="/login", status_code=303)

    return templates.TemplateResponse(
        "nfe_importar.html",
        {"request": request, "title": "Importar NF-e", "ok": ok, "error": error, "itens": []}
    )

@app.post("/nfe/importar", response_class=HTMLResponse)
async def nfe_importar_post(request: Request):
    if not current_user(request):
        return RedirectResponse(url="/login", status_code=303)

    try:
        require_role(request, {"admin", "recebimento"})
    except Exception as e:
        return RedirectResponse(url=f"/nfe/importar?error={str(e)}", status_code=303)

    form = await request.form()
    file = form.get("arquivo")
    if not file:
        return RedirectResponse(url="/nfe/importar?error=Envie%20o%20arquivo%20XML", status_code=303)

    xml_bytes = await file.read()

    try:
        itens = parse_nfe_xml(xml_bytes)
    except Exception as e:
        return RedirectResponse(url=f"/nfe/importar?error={str(e)}", status_code=303)

    return templates.TemplateResponse(
        "nfe_importar.html",
        {"request": request, "title": "Importar NF-e", "ok": "", "error": "", "itens": itens}
    )

@app.post("/nfe/confirmar")
async def nfe_confirmar_post(request: Request):
    if not current_user(request):
        return RedirectResponse(url="/login", status_code=303)

    try:
        require_role(request, {"admin", "recebimento"})
    except Exception as e:
        return RedirectResponse(url=f"/nfe/importar?error={str(e)}", status_code=303)

    form = await request.form()

    rua = (form.get("rua") or "").strip()
    predio = (form.get("predio") or "").strip()
    nivel = (form.get("nivel") or "").strip()
    apartamento = (form.get("apartamento") or "").strip()
    endereco = f"Rua {rua} • Prédio {predio} • Nível {nivel} • Ap {apartamento}"

    user = current_user(request)
    criado_por = user.get("username", "") if user else ""

    try:
        total = int(form.get("total_itens") or "0")
    except Exception:
        total = 0

    if total <= 0:
        return RedirectResponse(url="/nfe/importar?error=Nenhum%20item%20para%20confirmar", status_code=303)

    for i in range(total):
        codigo = (form.get(f"codigo_{i}") or "").strip()
        descricao = (form.get(f"descricao_{i}") or "").strip()
        lote = (form.get(f"lote_{i}") or "").strip()
        validade_raw = (form.get(f"validade_{i}") or "").strip()
        laboratorio = (form.get(f"laboratorio_{i}") or "").strip() or "N/I"

        try:
            quantidade = int(float((form.get(f"quantidade_{i}") or "0").replace(",", ".")))
        except Exception:
            quantidade = 0

        if not codigo or not descricao or quantidade <= 0:
            continue

        if not lote or not validade_raw:
            return RedirectResponse(url="/nfe/importar?error=Existem%20itens%20sem%20lote%20ou%20validade", status_code=303)

        try:
            validade = normalize_validade_ddmmyyyy(validade_raw)
        except Exception as e:
            return RedirectResponse(url=f"/nfe/importar?error={str(e)}", status_code=303)

        entrada_insert(codigo, descricao, lote, laboratorio, validade, quantidade, endereco, criado_por=criado_por)
        estoque_upsert(codigo, descricao, lote, laboratorio, validade, endereco, quantidade)

    return RedirectResponse(url="/entrada/novo?ok=1", status_code=303)

# ==========================
# ROTAS AUTH
# ==========================
@app.get("/login", response_class=HTMLResponse)
def login_get(request: Request, error: str = ""):
    return templates.TemplateResponse("login.html", {"request": request, "title": "Login", "error": error})

@app.post("/login")
def login_post(request: Request, username: str = Form(...), password: str = Form(...)):
    username = (username or "").strip()
    password = password or ""

    r = user_get(username)
    if not r:
        return RedirectResponse(url="/login?error=Usuário%20ou%20senha%20inválidos", status_code=303)

    u, ph, role = r
    if not verify_password(password, ph):
        return RedirectResponse(url="/login?error=Usuário%20ou%20senha%20inválidos", status_code=303)

    request.session["user"] = {"username": u, "role": role}
    return RedirectResponse(url="/", status_code=303)

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)

# ==========================
# ROTAS PRINCIPAIS
# ==========================
@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    if not current_user(request):
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("index.html", {"request": request, "title": "Início"})

@app.get("/entrada", response_class=HTMLResponse)
def entrada_menu(request: Request):
    if not current_user(request):
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("entrada_menu.html", {"request": request, "title": "Entrada"})

@app.get("/entrada/novo", response_class=HTMLResponse)
def entrada_get(request: Request, ok: int = 0, error: str = ""):
    if not current_user(request):
        return RedirectResponse(url="/login", status_code=303)

    return templates.TemplateResponse(
        "entrada.html",
        {"request": request, "title": "Entrada Manual", "ok": bool(ok), "error": error}
    )

@app.post("/entrada/novo")
def entrada_post(
    request: Request,
    codigo: str = Form(...),
    descricao: str = Form(...),
    lote: str = Form(...),
    laboratorio: str = Form(...),
    validade: str = Form(...),
    quantidade: int = Form(...),
    rua: str = Form(...),
    predio: str = Form(...),
    nivel: str = Form(...),
    apartamento: str = Form(...),
):
    if not current_user(request):
        return RedirectResponse(url="/login", status_code=303)

    try:
        require_role(request, {"admin", "recebimento"})
    except Exception as e:
        return RedirectResponse(url=f"/entrada/novo?error={str(e)}", status_code=303)

    try:
        quantidade = int(quantidade)
        if quantidade <= 0:
            raise ValueError("Quantidade inválida.")

        validade_padrao = normalize_validade_ddmmyyyy(validade)

        rua = (rua or "").strip()
        predio = (predio or "").strip()
        nivel = (nivel or "").strip()
        apartamento = (apartamento or "").strip()
        endereco = f"Rua {rua} • Prédio {predio} • Nível {nivel} • Ap {apartamento}"

        user = current_user(request)
        criado_por = user.get("username", "") if user else ""

        entrada_insert(
            codigo.strip(), descricao.strip(), lote.strip(),
            laboratorio.strip(), validade_padrao, quantidade, endereco,
            criado_por=criado_por
        )

        estoque_upsert(
            codigo.strip(), descricao.strip(), lote.strip(),
            laboratorio.strip(), validade_padrao, endereco, quantidade
        )

        return RedirectResponse(url="/entrada/novo?ok=1", status_code=303)
    except Exception as e:
        return RedirectResponse(url=f"/entrada/novo?error={str(e)}", status_code=303)

# ==========================
# RELATÓRIOS (HTML + CSV + XLSX)
# ==========================
@app.get("/relatorios/validade", response_class=HTMLResponse)
def relatorio_validade_get(request: Request, dias: int = 90):
    if not current_user(request):
        return RedirectResponse(url="/login", status_code=303)

    try:
        require_role(request, {"admin", "recebimento", "separador"})
    except Exception as e:
        return RedirectResponse(url=f"/?error={str(e)}", status_code=303)

    dias = int(dias)
    itens = relatorio_validade(dias)
    titulo = "Vencidos" if dias == 0 else f"Vencendo em até {dias} dias"

    return templates.TemplateResponse(
        "relatorio_validade.html",
        {
            "request": request,
            "title": "Relatório de Validade",
            "titulo": titulo,
            "dias": dias,
            "itens": itens,
        }
    )

@app.get("/relatorios/validade.csv")
def relatorio_validade_csv(request: Request, dias: int = 90):
    if not current_user(request):
        return RedirectResponse(url="/login", status_code=303)

    try:
        require_role(request, {"admin", "recebimento", "separador"})
    except Exception:
        return RedirectResponse(url="/", status_code=303)

    dias = int(dias)
    itens = relatorio_validade(dias)

    titulo = "vencidos" if dias == 0 else f"vence_{dias}_dias"
    filename = f"relatorio_validade_{titulo}.csv"

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["ID", "CODIGO", "DESCRICAO", "LOTE", "VALIDADE", "ENDERECO", "QTD", "DIAS_PARA_VENCER"])

    for r in itens:
        writer.writerow([r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7]])

    output.seek(0)
    resp = StreamingResponse(iter([output.getvalue()]), media_type="text/csv; charset=utf-8")
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

@app.get("/relatorios/validade.xlsx")
def relatorio_validade_xlsx(request: Request, dias: int = 90):
    if not current_user(request):
        return RedirectResponse(url="/login", status_code=303)

    try:
        require_role(request, {"admin", "recebimento", "separador"})
    except Exception:
        return RedirectResponse(url="/", status_code=303)

    dias = int(dias)
    itens = relatorio_validade(dias)

    titulo = "vencidos" if dias == 0 else f"vence_{dias}_dias"
    filename = f"relatorio_validade_{titulo}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Validade"

    header = ["ID", "CÓDIGO", "DESCRIÇÃO", "LOTE", "VALIDADE", "ENDEREÇO", "QTD", "DIAS"]
    ws.append(header)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2A44")
    header_align = Alignment(vertical="center")

    for col in range(1, len(header) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    for r in itens:
        ws.append([r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7]])

    widths = [8, 14, 40, 14, 14, 28, 8, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    fill_vencido = PatternFill("solid", fgColor="7F1D1D")
    fill_urgente = PatternFill("solid", fgColor="7C2D12")
    fill_ok = PatternFill("solid", fgColor="0F3D2E")

    for row in range(2, ws.max_row + 1):
        dias_cell = ws.cell(row=row, column=8)
        try:
            d = int(dias_cell.value)
        except Exception:
            continue

        if d < 0:
            dias_cell.fill = fill_vencido
            dias_cell.font = Font(bold=True, color="FFFFFF")
        elif d <= 30:
            dias_cell.fill = fill_urgente
            dias_cell.font = Font(bold=True, color="FFFFFF")
        else:
            dias_cell.fill = fill_ok
            dias_cell.font = Font(color="FFFFFF")

        dias_cell.alignment = Alignment(horizontal="center")

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_name = tmp.name
    tmp.close()
    wb.save(tmp_name)

    def file_iter():
        with open(tmp_name, "rb") as f:
            yield from f
        try:
            os.remove(tmp_name)
        except Exception:
            pass

    resp = StreamingResponse(file_iter(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

# ==========================
# USUÁRIOS (ADMIN)
# ==========================
@app.get("/usuarios", response_class=HTMLResponse)
def usuarios_get(request: Request, ok: str = "", error: str = ""):
    if not current_user(request):
        return RedirectResponse(url="/login", status_code=303)

    try:
        require_role(request, {"admin"})
    except Exception as e:
        return RedirectResponse(url=f"/?error={str(e)}", status_code=303)

    rows = users_list()
    return templates.TemplateResponse(
        "usuarios.html",
        {"request": request, "title": "Usuários", "rows": rows, "ok": ok, "error": error}
    )

@app.post("/usuarios/create")
def usuarios_create(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    role: str = Form(...),
):
    if not current_user(request):
        return RedirectResponse(url="/login", status_code=303)

    try:
        require_role(request, {"admin"})
    except Exception as e:
        return RedirectResponse(url=f"/usuarios?error={str(e)}", status_code=303)

    try:
        user_create(username, password, role)
        return RedirectResponse(url="/usuarios?ok=Usuário%20criado%20✅", status_code=303)
    except Exception as e:
        return RedirectResponse(url=f"/usuarios?error={str(e)}", status_code=303)

@app.post("/usuarios/delete/{user_id}")
def usuarios_delete(request: Request, user_id: int):
    if not current_user(request):
        return RedirectResponse(url="/login", status_code=303)

    try:
        require_role(request, {"admin"})
    except Exception as e:
        return RedirectResponse(url=f"/usuarios?error={str(e)}", status_code=303)

    try:
        u = current_user(request)
        current_username = (u.get("username") if u else "") or ""
        user_delete(int(user_id), current_username=current_username)
        return RedirectResponse(url="/usuarios?ok=Usuário%20excluído%20✅", status_code=303)
    except Exception as e:
        return RedirectResponse(url=f"/usuarios?error={str(e)}", status_code=303)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=int(os.getenv("PORT", "8000")), reload=True)